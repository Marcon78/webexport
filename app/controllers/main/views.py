#!/usr/bin/env python
# -*- coding:utf-8 -*-
import time
from sqlalchemy import create_engine, select, func, String, Integer, Table, Column, MetaData
from sqlalchemy.orm import scoped_session, sessionmaker
from flask import request, render_template, current_app, jsonify, make_response, send_file
from app import db

from . import main

@main.route("/")
def index():
    # engine = create_engine(current_app.config["SQLALCHEMY_DATABASE_URI"],
    #                     convert_unicode=True)
    # metadata = MetaData()
    # db_session = scoped_session(sessionmaker(autocommit=False,
    #                                         autoflush=False,
    #                                         bind=engine))

    # s = db_session()
    # res = s.execute("SELECT * FROM tbl_saleinfo;").fetchall()
    # res = db.session.execute(
    #     """SELECT
    #            (SELECT COUNT(*)
    #             FROM tbl_purchaseinfo AS t2
    #             WHERE t2.CommodityNo < t1.CommodityNo
    #            ) + (
    #            SELECT COUNT(*)
    #            FROM tbl_purchaseinfo AS t3
    #            WHERE t3.name = t1.name
    #              AND t3.id < t1.id
    #            ) + 1 AS rowNum,
    #            *
    #        FROM tbl_purchaseinfo AS t1
    #        ORDER BY t1.CommodityNo ASC
    #        LIMIT(50);""").fetchall();
    res = db.session.execute("SELECT '', * FROM tbl_purchaseinfo LIMIT(5000);").fetchall()
    return render_template("index.html", res=res)


@main.route("/data", methods=["GET", "POST"])
def data():
    metadata = MetaData(db.engine)
    # 如果仅仅是取数据条目数，即 count(*)，那么字段定义可以省略。
    table = Table("tbl_purchaseinfo", metadata
                  # ,
                  # Column('primary_key', Integer),
                  # Column('other_column', Integer)  # just to illustrate
                  )

    # 定义列名
    CONST_COLS = ["#", "CommodityNo", "CommodityName", "Quantity", "RQuantity", "Price", "MakeDate", "Maker"]

    ####
    # DataTables 发送到服务器的参数
    #
    # 绘制计数器。这个是用来确保Ajax从服务器返回的是对应的（Ajax是异步的，因此返回的顺序是不确定的）。 要求在服务器接收到此参数后再返回。
    front_dt_draw = request.form.get("draw")
    # 分页的第一条数据的起始位置，0代表第一条数据。
    front_dt_start = request.form.get("start")
    # 告诉服务器每页显示的条数，这个数字会等于返回的 data集合的记录数，可能会大于因为服务器可能没有那么多数据。这个也可能是-1，代表需要返回全部数据。
    front_dt_length = request.form.get("length")
    # 全局的搜索条件，条件会应用到每一列（searchable需要设置为true）。
    front_dt_search_val = request.form.get("search[value]")
    # 如果为 true代表全局搜索的值是作为正则表达式处理，为 false则不是。 注意：通常在服务器模式下对于大数据不执行这样的正则表达式，但这都是自己决定的。
    front_dt_search_regx = request.form.get("search[regex]")
    front_dt_column_cnt = request.form.get("col_cnt")
    # order[i][column]
    # order[i][dir]

    ####
    # 服务器需要返回的数据
    #
    filter_params = ""
    order_params = ""
    for x in range(int(front_dt_column_cnt)):
        ord_idx = request.form.get("order["+str(x)+"][column]")
        ord_dir = request.form.get("order["+str(x)+"][dir]")
        if ord_idx and ord_dir:
            order_params += CONST_COLS[int(ord_idx)] + " " + ord_dir + ","
    if order_params:
        order_params = order_params[:-1]
    sql_txt = """SELECT '', *
                 FROM tbl_purchaseinfo """\
              + ("""ORDER BY """ + order_params if order_params else """""")\
              + """ LIMIT(""" + front_dt_length + """)
                   OFFSET """ + front_dt_start + """;"""
    res = db.session.execute(sql_txt).fetchall()
    back_dt = {
        # 必要！Datatables发送的draw是多少，则服务器返回多少。
        # 注意！！！出于安全的考虑，强烈要求把这个转换为整形，即数字后再返回，而不是纯粹的接受然后返回，目的是防止跨站脚本（XSS）攻击。
        "draw": int(front_dt_draw),
        # 必要！即没有过滤的记录数（数据库里总共记录数）。
        "recordsTotal": select([func.count()]).select_from(table).scalar()
    }
    # 必要！过滤后的记录数（如果有接收到前台的过滤条件，则返回的是过滤后的记录数）。
    back_dt["recordsFiltered"] = back_dt["recordsTotal"]
    # 必要！表中中需要显示的数据。这是一个对象数组，也可以只是数组。
    # 区别在于：纯数组前台无需用 columns 绑定数据，会自动按照顺序去显示；对象数组则需要用 columns 绑定数据才能正常显示。
    back_dt_data = []
    for r in res:
        back_dt_data.append(list(r))
    back_dt["data"] = back_dt_data
    # 可选。你可以定义一个错误来描述服务器出了问题后的友好提示。
    # back_dt["error"] = ""
    # 自动绑定Id到tr节点上。string。
    # DT_RowId
    # 自动把这个类名添加到 tr。string。
    # DT_RowClass
    # 使用 jQuery.data() 方法把数据绑定到row中，方便之后用来检索（比如加入一个点击事件）。object。
    # DT_RowData
    # 自动绑定数据到 tr上，使用 jQuery.attr() 方法，对象的键用作属性，值用作属性的值。注意这个 需要 Datatables 1.10.5+的版本才支持。object。
    # DT_RowAttr

    # jsonify的作用实际上就是将传入的json形式数据序列化成为json字符串，作为响应的body，并且设置响应的Content-Type为application/json，构造出响应返回至客户端。
    # 直接返回json.dumps的结果是可行的，因为flask会判断并使用make_response方法自动构造出响应，只不过响应头各个字段是默认的。
    # 若要自定义响应字段，则可以使用make_response或Response自行构造响应。
    return jsonify(back_dt)


from openpyxl.styles import Border, Side, Color, Fill, PatternFill, GradientFill, Font, Alignment
def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

@main.route("/export1", methods=["GET"])
def export1():
    t0 = time.clock()

    # https://bitbucket.org/openpyxl/openpyxl/src
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook
    # # from openpyxl.writer.excel import ExcelWriter
    # from openpyxl.styles import Color, Fill, PatternFill, GradientFill, Font, Alignment
    # # from openpyxl.cell import Cell
    # # from openpyxl.utils import get_column_letter

    # 新建一个workbook
    wb = Workbook(write_only=True)
    # 第一个sheet是ws
    ws = wb.create_sheet()
    # 设置ws的名称
    ws.title = "统计"
    ws.append(["CommodityNo", "CommodityName", "Quantity", "RealQuantity", "Price", "MakeDate", "Maker"])
    # thin = Side(border_style="thin", color="000000")
    # double = Side(border_style="double", color="ff0000")
    # border = Border(top=double, left=thin, right=thin, bottom=double)
    # fill = PatternFill("solid", fgColor="DDDDDD")
    # # fill2 = GradientFill(stop=("000000", "FFFFFF"))
    # font = Font(b=True, color="FF0000")
    # al = Alignment(horizontal="center", vertical="center")
    # # write_only 时不允许进行子脚本操作，也就无法定义相关的单元样式。
    # style_range(ws, 'B2:F4', border=border, fill=fill, font=font, alignment=al)

    # # 给A1赋值
    # ws['A1'].value = '%s' % ("跟随总数")
    # # 给A2赋值
    # # 先把数字转换成字母
    # col = get_column_letter(1)
    # # 赋值
    # ws.cell('%s%s' % (col, 2)).value = '%s' % ("A2")
    # # 字体修改样式
    # ##颜色
    # ws.cell('A2').style.font.color.index = Color.GREEN
    # ##字体名称
    # ws.cell('A2').style.font.name = 'Arial'
    # ##字号
    # ws.cell('A2').style.font.size = 8
    # ##加粗
    # ws.cell('A2').style.font.bold = True
    # ##不知道干啥用的
    # ws.cell('A2').style.alignment.wrap_text = True
    # ##背景 好像不太好用 是个BUG
    # ws.cell('A2').style.fill.fill_type = Fill.FILL_SOLID
    # ws.cell('A2').style.fill.start_color.index = Color.DARKRED
    # ##修改某一列宽度
    # ws.column_dimensions["C"].width = 60.0
    # ##增加一个表
    # ws = wb.create_sheet()
    # ws.title = u'结单统计'
    # # # 保存生成xlsx
    # # file_name = 'test.xlsx'
    # # file_dir = '/home/x/'
    # # dest_filename = '%s%s' % (file_dir, file_name)
    # # ew = ExcelWriter(workbook=wb)
    # ew = ExcelWriter(workbook=wb)
    res = db.session.execute("SELECT * FROM tbl_purchaseinfo;").fetchall()
    for r in res:
        ws.append(list(r))

    # wb.save("new_big_file.xlsx")
    # ew = ExcelWriter(workbook=wb)
    # response = make_response(send_file("..//new_big_file.xlsx"))
    response = make_response(save_virtual_workbook(wb))
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = "attachment; filename=new_big_file1.xlsx;"

    t = time.clock() - t0
    print(t)

    return response


@main.route("/export2", methods=["GET"])
def export2():
    t0 = time.clock()

    # https://github.com/pyexcel/pyexcel
    import pyexcel as pe

    data = [["CommodityNo", "CommodityName", "Quantity", "RealQuantity", "Price", "MakeDate", "Maker"]]
    res = db.session.execute("SELECT * FROM tbl_purchaseinfo;").fetchall()
    for r in res:
        data.append(list(r))
    sheet = pe.Sheet(data)
    response = make_response(sheet.xlsx)
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = "attachment; filename=new_big_file2.xlsx;"

    t = time.clock() - t0
    print(t)

    return response


# quick
@main.route("/export3", methods=["GET"])
def export3():
    t0 = time.clock()

    # xlrd/xlwt 只对xls的格式处理的比较好。
    # https://github.com/python-excel/xlwt
    import xlwt
    from io import BytesIO

    wb = xlwt.Workbook()
    wb.encoding ='gbk'
    ws = wb.add_sheet('1')
    titles = ["CommodityNo", "CommodityName", "Quantity", "RealQuantity", "Price", "MakeDate", "Maker"]
    for i in range(len(titles)):
        ws.write(0, i, titles[i])

    res = db.session.execute("SELECT * FROM tbl_purchaseinfo;").fetchall()
    row = 1
    for r in res:
        for i in range(len(titles)):
            ws.write(row, i, list(r)[i])
        row += 1

    bio = BytesIO()
    wb.save(bio)   #这点很重要，传给save函数的不是保存文件名，而是一个StringIO流
    response = make_response(bio.getvalue())
    response.headers["Content-type"] = "application/vnd.ms-excel"   # 指定返回的类型
    # response.headers["Transfer-Encoding"] = "chunked"
    response.headers["Content-Disposition"] = "attachment;filename=new_big_file3.xls"  # 设定用户浏览器显示的保存文件名

    t = time.clock() - t0
    print(t)
    return response


@main.route("/export4", methods=["GET"])
def export4():
    t0 = time.clock()

    # https://github.com/kennethreitz/tablib
    import tablib
    titles = ["CommodityNo", "CommodityName", "Quantity", "RealQuantity", "Price", "MakeDate", "Maker"]
    res = db.session.execute("SELECT * FROM tbl_purchaseinfo;").fetchall()
    data = tablib.Dataset(*list(res), headers=titles)
    response = make_response(data.xlsx)
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = "attachment; filename=new_big_file2.xlsx;"

    t = time.clock() - t0
    print(t)

    return response